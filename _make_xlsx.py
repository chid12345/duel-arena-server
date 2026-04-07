"""Генерация progression_100_levels_v4.xlsx — актуальная таблица v7."""
import json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── данные ──────────────────────────────────────────────────────────────────
p = json.load(open("progression_100_levels_v4/progression.json", encoding="utf-8"))
XP_NEED   = p["xp_to_next"]       # len 99
XP_WIN    = p["xp_per_win"]       # len 100
STEPS     = p["steps_per_level"]  # len 99
STATS_ON  = p["stats_on_reach"]   # len 101
HP_ON     = p["hp_on_reach"]      # len 101
GOLD_ON   = p["gold_on_reach"]    # len 101

WIN_RATE  = 0.42   # среднестат. винрейт
MIN_PER_BATTLE = 2.0  # минут на бой с очередью
PREM_MULT = 1.30

ZONES = {
    (1,  4):  ("Новичок",  "4472C4"),
    (5,  10): ("Боец",     "70AD47"),
    (11, 22): ("Солдат",   "FFC000"),
    (23, 40): ("Ветеран",  "FF7F00"),
    (41, 65): ("Элита",    "C00000"),
    (66, 85): ("Легенда",  "7030A0"),
    (86, 100):("Мастер",   "1F1F1F"),
}
SPECIAL = {
    5:  "Разблокировка рейтинга новичков",
    10: "Звание «Солдат» + ранговые бои",
    20: "Открывается лига Ветеранов",
    40: "Звание «Ветеран»",
    50: "Открывается Элитная лига",
    65: "Звание «Элита»",
    75: "Открывается лига Легенд",
    85: "Звание «Легенда»",
    99: "Последний рывок к мастерству",
    100:"🏆 МАСТЕР — максимальный уровень",
}

def zone_label(lv):
    for (lo, hi), (name, _) in ZONES.items():
        if lo <= lv <= hi:
            return name
    return ""

def zone_color(lv):
    for (lo, hi), (name, color) in ZONES.items():
        if lo <= lv <= hi:
            return color
    return "FFFFFF"

# ── стили ───────────────────────────────────────────────────────────────────
def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
def border(): return Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FILL  = fill("1F3864")   # тёмно-синий
HDR_FONT  = font(bold=True, color="FFFFFF", size=10)
SUB_FILL  = fill("2E75B6")
SUB_FONT  = font(bold=True, color="FFFFFF", size=9)
EVEN_FILL = fill("EBF3FB")
ODD_FILL  = fill("FFFFFF")
PREM_FILL = fill("FFF2CC")   # жёлтый — колонки с подпиской
SPEC_FONT = font(bold=True, color="C00000", size=9)

# ── вспомогательные ─────────────────────────────────────────────────────────
def ap_thresholds(need, steps):
    return [(need * k) // (steps + 1) for k in range(1, steps + 1)]

def wins_to(xp, avg_xp_win):
    return xp / avg_xp_win if avg_xp_win > 0 else 0

def battles_to(xp, avg_xp_win):
    return wins_to(xp, avg_xp_win) / WIN_RATE if WIN_RATE > 0 else 0

def fmt_time(battles):
    mins = battles * MIN_PER_BATTLE
    h = int(mins // 60)
    m = int(mins % 60)
    if h == 0: return f"{m} мин"
    return f"{h}ч {m}м"

# ── накопительное XP ─────────────────────────────────────────────────────────
cum_xp = [0] * 101   # cum_xp[lv] = суммарный XP чтобы достичь lv (0 на старте)
for i in range(99):
    cum_xp[i + 2] = cum_xp[i + 1] + XP_NEED[i]

# ════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Лист 1: Сводная (всё) ───────────────────────────────────────────────────
ws = wb.active
ws.title = "Все уровни"

COLS = [
    # (заголовок, ширина)
    ("Ур",                 5),
    ("Зона",              11),
    ("XP на ур.",          9),
    ("Апов",               6),
    ("AP пороги (% полоски)", 32),
    ("Win XP",             8),
    ("Побед до 1-го апа",  11),
    ("Побед до ур.",       11),
    ("Боёв до ур.",        11),
    ("Время до ур.",       11),
    ("★ Побед до ур.",     12),  # premium
    ("★ Боёв до ур.",      12),
    ("★ Время до ур.",     12),
    ("+Стат при ур.",      10),
    ("+HP при ур.",         9),
    ("+Золото при ур.",    11),
    ("Нак. XP",            12),
    ("Нак. время",         11),
    ("Особое событие",     30),
]

# заголовок
ws.merge_cells("A1:S1")
c = ws["A1"]
c.value = "СИСТЕМА ПРОКАЧКИ — ВСЕ 100 УРОВНЕЙ"
c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
c.fill = HDR_FILL
c.alignment = center()
ws.row_dimensions[1].height = 24

# подзаголовок
ws.merge_cells("A2:S2")
c = ws["A2"]
c.value = (f"XP за победу: 85→102 по уровням  ·  С подпиской: ×{PREM_MULT}  ·  "
           f"Win rate: {int(WIN_RATE*100)}%  ·  ~{MIN_PER_BATTLE:.0f} мин на бой")
c.font = font(bold=False, color="FFFFFF", size=9)
c.fill = SUB_FILL
c.alignment = center()
ws.row_dimensions[2].height = 16

# шапка колонок
for ci, (title, width) in enumerate(COLS, start=1):
    cl = get_column_letter(ci)
    ws.column_dimensions[cl].width = width
    c = ws.cell(row=3, column=ci, value=title)
    c.font = SUB_FONT
    c.fill = HDR_FILL
    c.alignment = center()
    c.border = border()
ws.row_dimensions[3].height = 36

# пометим premium-колонки
PREM_COL_IDX = [11, 12, 13]  # 1-based: 11,12,13 → ★-колонки
for ci in PREM_COL_IDX:
    ws.cell(row=3, column=ci).fill = fill("7030A0")

# данные
cum_win_normal = 0.0
cum_win_prem   = 0.0
cum_bat_normal = 0.0
cum_bat_prem   = 0.0

for lv in range(1, 101):
    row = lv + 3
    is_even = lv % 2 == 0
    base_fill = EVEN_FILL if is_even else ODD_FILL

    if lv <= 99:
        need  = XP_NEED[lv - 1]
        steps = STEPS[lv - 1]
    else:
        need  = 0
        steps = 0

    win_xp      = XP_WIN[lv - 1]
    win_xp_prem = win_xp * PREM_MULT

    # апы
    thrs = ap_thresholds(need, steps) if need > 0 else []
    pcts = " · ".join(f"{int(100*t/need)}%" for t in thrs) if thrs and need > 0 else "—"
    first_ap_xp = thrs[0] if thrs else 0
    wins_first_ap = wins_to(first_ap_xp, win_xp) if first_ap_xp > 0 else 0

    # победы/бои до следующего уровня
    w_normal = wins_to(need, win_xp)         if need > 0 else 0
    b_normal = battles_to(need, win_xp)      if need > 0 else 0
    w_prem   = wins_to(need, win_xp_prem)    if need > 0 else 0
    b_prem   = battles_to(need, win_xp_prem) if need > 0 else 0

    cum_win_normal += w_normal
    cum_win_prem   += w_prem
    cum_bat_normal += b_normal
    cum_bat_prem   += b_prem

    stat  = STATS_ON[lv] if lv < len(STATS_ON) else 0
    hp    = HP_ON[lv]    if lv < len(HP_ON)    else 0
    gold  = GOLD_ON[lv]  if lv < len(GOLD_ON)  else 0
    special = SPECIAL.get(lv, "")

    row_data = [
        lv,
        zone_label(lv),
        need if need > 0 else "—",
        steps if steps > 0 else "—",
        pcts,
        win_xp,
        round(wins_first_ap, 1) if wins_first_ap > 0 else "—",
        round(w_normal, 1) if w_normal > 0 else "—",
        round(b_normal, 1) if b_normal > 0 else "—",
        fmt_time(b_normal) if b_normal > 0 else "—",
        round(w_prem, 1) if w_prem > 0 else "—",
        round(b_prem, 1) if b_prem > 0 else "—",
        fmt_time(b_prem) if b_prem > 0 else "—",
        stat,
        hp,
        gold,
        cum_xp[lv],
        fmt_time(cum_bat_normal),
        special,
    ]

    zcolor = zone_color(lv)
    for ci, val in enumerate(row_data, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.border = border()
        c.alignment = center() if ci != 5 and ci != 19 else left()
        # зональный цвет для колонки Ур и Зона
        if ci in (1, 2):
            c.fill = fill(zcolor)
            c.font = font(bold=(ci == 1), color="FFFFFF" if zcolor != "FFFFFF" else "000000", size=10)
        elif ci in PREM_COL_IDX:
            c.fill = PREM_FILL
            c.font = font(size=9)
        elif special and ci == 19:
            c.fill = fill("FFE0E0")
            c.font = SPEC_FONT
        else:
            c.fill = base_fill
            c.font = font(size=9)

    ws.row_dimensions[row].height = 16

ws.freeze_panes = "A4"

# ── Лист 2: Легенда ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Легенда")
legend = [
    ("ЛЕГЕНДА К ТАБЛИЦЕ", "", "1F3864", True),
    ("", "", "", False),
    ("Колонка", "Описание", "2E75B6", True),
    ("Ур",             "Номер уровня (1–100)", "", False),
    ("Зона",           "Тир персонажа (Новичок → Мастер)", "", False),
    ("XP на ур.",      "Сколько XP нужно на полоске этого уровня до следующего", "", False),
    ("Апов",           "Сколько промежуточных +1 стат выдаётся на полоске", "", False),
    ("AP пороги",      "На каком % полоски стоит каждый ап (равномерно)", "", False),
    ("Win XP",         "Базовый XP за победу на этом уровне (85 на ур.1 → 102 на ур.100)", "", False),
    ("Побед до 1-го апа", "Побед чтобы добраться до первого апа на полоске", "", False),
    ("Побед до ур.",   "Побед чтобы пройти всю полоску (выиграть следующий уровень)", "", False),
    ("Боёв до ур.",    "Боёв (победа + поражение) при win rate 42%", "", False),
    ("Время до ур.",   "Примерное время на уровень при 2 мин/бой", "", False),
    ("★ — колонки",   "То же самое с Premium подпиской (+30% XP за каждый бой)", "", False),
    ("+Стат при ур.",  "Свободных статов при достижении нового уровня", "", False),
    ("+HP при ур.",    "Прирост max HP при достижении нового уровня", "", False),
    ("+Золото при ур.","Золото при достижении нового уровня", "", False),
    ("Нак. XP",        "Суммарный XP с ур.1 до начала этого уровня", "", False),
    ("Нак. время",     "Суммарное время с ур.1 (без подписки, 42% WR, 2 мин/бой)", "", False),
    ("", "", "", False),
    ("ЗОНЫ", "", "2E75B6", True),
]
for z, (lo, hi), (name, color) in [
    (1, (1,4),   ("Новичок",   "4472C4")),
    (2, (5,10),  ("Боец",      "70AD47")),
    (3, (11,22), ("Солдат",    "FFC000")),
    (4, (23,40), ("Ветеран",   "FF7F00")),
    (5, (41,65), ("Элита",     "C00000")),
    (6, (66,85), ("Легенда",   "7030A0")),
    (7, (86,100),("Мастер",    "1F1F1F")),
]:
    legend.append((f"Ур. {lo}–{hi}", name, color, False))

for r, (a, b, fc, bold) in enumerate(legend, start=1):
    ca = ws2.cell(row=r, column=1, value=a)
    cb = ws2.cell(row=r, column=2, value=b)
    if fc:
        ca.fill = fill(fc)
        cb.fill = fill(fc)
        ca.font = font(bold=bold, color="FFFFFF", size=10)
        cb.font = font(bold=bold, color="FFFFFF", size=10)
    else:
        ca.font = font(bold=False, size=9)
        cb.font = font(bold=False, size=9)
    ca.alignment = left()
    cb.alignment = left()
    ca.border = border()
    cb.border = border()
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 52

# ── сохранить ────────────────────────────────────────────────────────────────
out = "progression_100_levels_v4/progression_100_levels_v4.xlsx"
wb.save(out)
print("Saved:", out)
