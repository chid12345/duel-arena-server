"""
tools/_balance_xlsx_writer.py — вывод payload в xlsx-файл.

Внутренний модуль balance_xlsx_export. Отдельный «дом» по Закону 2 — вся
xlsx-логика тут, генератор кривых не знает про openpyxl.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_HDR_FILL = PatternFill("solid", start_color="FFD966")
_HDR_FONT = Font(bold=True)
_NOTE_FONT = Font(italic=True, color="666666")
_INPUT_FONT = Font(color="0000FF", bold=True)


def _hdr(ws, row: int, values: list[str]) -> None:
    for col, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def _write_anchor(wb: Workbook, payload: dict[str, Any]) -> None:
    a = wb.create_sheet("Anchor")
    a["A1"] = "ЯКОРИ Duel Arena (генерируются CONFIG в tools/balance_xlsx_export.py)"
    a["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Дней до 80 уровня (средний игрок)", payload["anchor"]["days_to_max_level"]),
        ("PU/день (часов активной игры)",       payload["anchor"]["pu_per_day"]),
        ("Максимальный уровень",                 payload["anchor"]["max_level"]),
        ("XP за 1 PU (среднее, калиброванное)",  payload["anchor"]["xp_per_pu_avg"]),
        ("Tier T2 разблокирован с уровня",       payload["tier_thresholds"]["T2"]),
        ("Tier T3 разблокирован с уровня",       payload["tier_thresholds"]["T3"]),
        ("Tier T4 разблокирован с уровня",       payload["tier_thresholds"]["T4"]),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        a.cell(row=i, column=1, value=k)
        a.cell(row=i, column=2, value=v).font = _INPUT_FONT
    a.column_dimensions["A"].width = 50
    a.column_dimensions["B"].width = 14


def _write_curves(wb: Workbook, payload: dict[str, Any]) -> None:
    c = wb.create_sheet("Curves_by_level")
    _hdr(c, 1, ["Уровень", "XP до след.", "XP суммарно", "Мощь", "Дней до уровня",
                "Tier-unlock", "Доступные тиры", "PvP брекет", "Золото/PU"])
    for i, r in enumerate(payload["by_level"], start=2):
        c.cell(row=i, column=1, value=r["level"])
        c.cell(row=i, column=2, value=r["xp_to_next"])
        c.cell(row=i, column=3, value=r["xp_cum"])
        c.cell(row=i, column=4, value=r["power"])
        c.cell(row=i, column=5, value=r["days_to_reach"])
        c.cell(row=i, column=6, value=r["tier_unlock"])
        c.cell(row=i, column=7, value=",".join(r["tiers_available"]))
        c.cell(row=i, column=8, value=r["pvp_bracket"])
        c.cell(row=i, column=9, value=r["gold_per_pu"])
    for col in "ABCDEFGHI":
        c.column_dimensions[col].width = 16


def _write_brackets(wb: Workbook, payload: dict[str, Any]) -> None:
    p = wb.create_sheet("PvP_brackets")
    _hdr(p, 1, ["ID", "Уровень min", "Уровень max", "База XP за победу", "База золота за победу"])
    for i, b in enumerate(payload["pvp_brackets"], start=2):
        p.cell(row=i, column=1, value=b["id"])
        p.cell(row=i, column=2, value=b["min"])
        p.cell(row=i, column=3, value=b["max"])
        p.cell(row=i, column=4, value=b["xp_base"])
        p.cell(row=i, column=5, value=b["gold_base"])
    p["A8"] = "Внутри брекета — сила решает. Между брекетами — матчмейкинг не сводит."
    p["A8"].font = _NOTE_FONT


def _write_premium(wb: Workbook, payload: dict[str, Any]) -> None:
    pe = wb.create_sheet("Premium_effects")
    _hdr(pe, 1, ["Эффект", "Значение"])
    for i, (k, v) in enumerate(payload["premium_effects"].items(), start=2):
        pe.cell(row=i, column=1, value=k)
        pe.cell(row=i, column=2, value=v)
    pe["A8"] = "Премиум = только время и удобство. Никакого P2W-шмота."
    pe["A8"].font = _NOTE_FONT
    pe.column_dimensions["A"].width = 32


def _write_upgrades(wb: Workbook, payload: dict[str, Any]) -> None:
    u = wb.create_sheet("Upgrades_curve")
    _hdr(u, 1, ["Tier", "Max +N", "Прирост стат/шаг (%)", "Шанс провала с"])
    for i, (t, mx) in enumerate(payload["upgrades"]["max_plus_per_tier"].items(), start=2):
        u.cell(row=i, column=1, value=t)
        u.cell(row=i, column=2, value=mx)
        u.cell(row=i, column=3, value=payload["upgrades"]["stat_step_pct"] * 100)
        u.cell(row=i, column=4, value=f"+{payload['upgrades']['fail_chance_start']}")


def _write_sets(wb: Workbook, payload: dict[str, Any]) -> None:
    s = wb.create_sheet("Sets_catalog")
    _hdr(s, 1, ["ID", "Название", "Эмодзи", "Порог 3", "Порог 5", "Порог 7"])
    for i, st in enumerate(payload["sets"], start=2):
        s.cell(row=i, column=1, value=st["id"])
        s.cell(row=i, column=2, value=st["name"])
        s.cell(row=i, column=3, value=st["emoji"])
        s.cell(row=i, column=4, value="бонус +%")
        s.cell(row=i, column=5, value="усиленный бонус")
        s.cell(row=i, column=6, value="перк (только при полном составе)")
    s.column_dimensions["B"].width = 18
    s.column_dimensions["F"].width = 38


def write_xlsx(payload: dict[str, Any], out_path: Path) -> None:
    """Сериализовать payload в xlsx-файл с 6 листами."""
    wb = Workbook()
    wb.remove(wb.active)
    _write_anchor(wb, payload)
    _write_curves(wb, payload)
    _write_brackets(wb, payload)
    _write_premium(wb, payload)
    _write_upgrades(wb, payload)
    _write_sets(wb, payload)
    wb.save(out_path)
