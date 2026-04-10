"""Лист «Все уровни»."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from _make_xlsx import data as D
from _make_xlsx.styles import (
    EVEN_FILL,
    HDR_FILL,
    HDR_FONT,
    ODD_FILL,
    PREM_FILL,
    SPEC_FONT,
    SUB_FILL,
    SUB_FONT,
    border,
    center,
    fill,
    font,
    left,
)

COLS = [
    ("Ур", 5),
    ("Зона", 11),
    ("XP на ур.", 9),
    ("Апов", 6),
    ("AP пороги (% полоски)", 32),
    ("Win XP", 8),
    ("Побед до 1-го апа", 11),
    ("Побед до ур.", 11),
    ("Боёв до ур.", 11),
    ("Время до ур.", 11),
    ("★ Побед до ур.", 12),
    ("★ Боёв до ур.", 12),
    ("★ Время до ур.", 12),
    ("+Стат при ур.", 10),
    ("+HP при ур.", 9),
    ("+Золото при ур.", 11),
    ("Нак. XP", 12),
    ("Нак. время", 11),
    ("Особое событие", 30),
]

PREM_COL_IDX = [11, 12, 13]


def build_all_levels_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Все уровни"

    ws.merge_cells("A1:S1")
    c = ws["A1"]
    c.value = "СИСТЕМА ПРОКАЧКИ — ВСЕ 100 УРОВНЕЙ"
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.fill = HDR_FILL
    c.alignment = center()
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:S2")
    c = ws["A2"]
    c.value = (
        f"XP за победу: 85→102 по уровням  ·  С подпиской: ×{D.PREM_MULT}  ·  "
        f"Win rate: {int(D.WIN_RATE * 100)}%  ·  ~{D.MIN_PER_BATTLE:.0f} мин на бой"
    )
    c.font = font(bold=False, color="FFFFFF", size=9)
    c.fill = SUB_FILL
    c.alignment = center()
    ws.row_dimensions[2].height = 16

    for ci, (title, width) in enumerate(COLS, start=1):
        cl = get_column_letter(ci)
        ws.column_dimensions[cl].width = width
        c = ws.cell(row=3, column=ci, value=title)
        c.font = SUB_FONT
        c.fill = HDR_FILL
        c.alignment = center()
        c.border = border()
    ws.row_dimensions[3].height = 36

    for ci in PREM_COL_IDX:
        ws.cell(row=3, column=ci).fill = fill("7030A0")

    cum_win_normal = 0.0
    cum_win_prem = 0.0
    cum_bat_normal = 0.0
    cum_bat_prem = 0.0

    for lv in range(1, 101):
        row = lv + 3
        is_even = lv % 2 == 0
        base_fill = EVEN_FILL if is_even else ODD_FILL

        if lv <= 99:
            need = D.XP_NEED[lv - 1]
            steps = D.STEPS[lv - 1]
        else:
            need = 0
            steps = 0

        win_xp = D.XP_WIN[lv - 1]
        win_xp_prem = win_xp * D.PREM_MULT

        thrs = D.ap_thresholds(need, steps) if need > 0 else []
        pcts = " · ".join(f"{int(100 * t / need)}%" for t in thrs) if thrs and need > 0 else "—"
        first_ap_xp = thrs[0] if thrs else 0
        wins_first_ap = D.wins_to(first_ap_xp, win_xp) if first_ap_xp > 0 else 0

        w_normal = D.wins_to(need, win_xp) if need > 0 else 0
        b_normal = D.battles_to(need, win_xp) if need > 0 else 0
        w_prem = D.wins_to(need, win_xp_prem) if need > 0 else 0
        b_prem = D.battles_to(need, win_xp_prem) if need > 0 else 0

        cum_win_normal += w_normal
        cum_win_prem += w_prem
        cum_bat_normal += b_normal
        cum_bat_prem += b_prem

        stat = D.STATS_ON[lv] if lv < len(D.STATS_ON) else 0
        hp = D.HP_ON[lv] if lv < len(D.HP_ON) else 0
        gold = D.GOLD_ON[lv] if lv < len(D.GOLD_ON) else 0
        special = D.SPECIAL.get(lv, "")

        row_data = [
            lv,
            D.zone_label(lv),
            need if need > 0 else "—",
            steps if steps > 0 else "—",
            pcts,
            win_xp,
            round(wins_first_ap, 1) if wins_first_ap > 0 else "—",
            round(w_normal, 1) if w_normal > 0 else "—",
            round(b_normal, 1) if b_normal > 0 else "—",
            D.fmt_time(b_normal) if b_normal > 0 else "—",
            round(w_prem, 1) if w_prem > 0 else "—",
            round(b_prem, 1) if b_prem > 0 else "—",
            D.fmt_time(b_prem) if b_prem > 0 else "—",
            stat,
            hp,
            gold,
            D.cum_xp[lv],
            D.fmt_time(cum_bat_normal),
            special,
        ]

        zcolor = D.zone_color(lv)
        for ci, val in enumerate(row_data, start=1):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = border()
            c.alignment = center() if ci != 5 and ci != 19 else left()
            if ci in (1, 2):
                c.fill = fill(zcolor)
                c.font = font(bold=(ci == 1), color="FFFFFF" if zcolor != "FFFFFF" else "000000", size=10)
            elif ci in PREM_COL_IDX:
                c.fill = PREM_FILL
                c.font = font(size=9)
            elif special and ci == 19:
                c.fill = fill("FFE0E0")
                c.font = SPEC_FONT
            else:
                c.fill = base_fill
                c.font = font(size=9)

        ws.row_dimensions[row].height = 16

    ws.freeze_panes = "A4"
    return ws
