"""Лист «Легенда»."""

from __future__ import annotations

from openpyxl import Workbook

from _make_xlsx.styles import border, fill, font, left


def build_legend_sheet(wb: Workbook):
    ws2 = wb.create_sheet("Легенда")
    legend = [
        ("ЛЕГЕНДА К ТАБЛИЦЕ", "", "1F3864", True),
        ("", "", "", False),
        ("Колонка", "Описание", "2E75B6", True),
        ("Ур", "Номер уровня (1–100)", "", False),
        ("Зона", "Тир персонажа (Новичок → Мастер)", "", False),
        ("XP на ур.", "Сколько XP нужно на полоске этого уровня до следующего", "", False),
        ("Апов", "Сколько промежуточных +1 стат выдаётся на полоске", "", False),
        ("AP пороги", "На каком % полоски стоит каждый ап (равномерно)", "", False),
        ("Win XP", "Базовый XP за победу на этом уровне (85 на ур.1 → 102 на ур.100)", "", False),
        ("Побед до 1-го апа", "Побед чтобы добраться до первого апа на полоске", "", False),
        ("Побед до ур.", "Побед чтобы пройти всю полоску (выиграть следующий уровень)", "", False),
        ("Боёв до ур.", "Боёв (победа + поражение) при win rate 42%", "", False),
        ("Время до ур.", "Примерное время на уровень при 2 мин/бой", "", False),
        ("★ — колонки", "То же самое с Premium подпиской (+30% XP за каждый бой)", "", False),
        ("+Стат при ур.", "Свободных статов при достижении нового уровня", "", False),
        ("+HP при ур.", "Прирост max HP при достижении нового уровня", "", False),
        ("+Золото при ур.", "Золото при достижении нового уровня", "", False),
        ("Нак. XP", "Суммарный XP с ур.1 до начала этого уровня", "", False),
        ("Нак. время", "Суммарное время с ур.1 (без подписки, 42% WR, 2 мин/бой)", "", False),
        ("", "", "", False),
        ("ЗОНЫ", "", "2E75B6", True),
    ]
    for z, (lo, hi), (name, color) in [
        (1, (1, 4), ("Новичок", "4472C4")),
        (2, (5, 10), ("Боец", "70AD47")),
        (3, (11, 22), ("Солдат", "FFC000")),
        (4, (23, 40), ("Ветеран", "FF7F00")),
        (5, (41, 65), ("Элита", "C00000")),
        (6, (66, 85), ("Легенда", "7030A0")),
        (7, (86, 100), ("Мастер", "1F1F1F")),
    ]:
        legend.append((f"Ур. {lo}–{hi}", name, color, False))

    for r, (a, b, fc, bold) in enumerate(legend, start=1):
        ca = ws2.cell(row=r, column=1, value=a)
        cb = ws2.cell(row=r, column=2, value=b)
        if fc:
            ca.fill = fill(fc)
            cb.fill = fill(fc)
            ca.font = font(bold=bold, color="FFFFFF", size=10)
            cb.font = font(bold=bold, color="FFFFFF", size=10)
        else:
            ca.font = font(bold=False, size=9)
            cb.font = font(bold=False, size=9)
        ca.alignment = left()
        cb.alignment = left()
        ca.border = border()
        cb.border = border()
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 52
    return ws2
