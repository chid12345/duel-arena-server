"""Стили openpyxl для таблицы прогрессии."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


thin = Side(style="thin", color="BFBFBF")


def border() -> Border:
    return Border(left=thin, right=thin, top=thin, bottom=thin)


HDR_FILL = fill("1F3864")
HDR_FONT = font(bold=True, color="FFFFFF", size=10)
SUB_FILL = fill("2E75B6")
SUB_FONT = font(bold=True, color="FFFFFF", size=9)
EVEN_FILL = fill("EBF3FB")
ODD_FILL = fill("FFFFFF")
PREM_FILL = fill("FFF2CC")
SPEC_FONT = font(bold=True, color="C00000", size=9)
